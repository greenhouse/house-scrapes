#!/usr/bin/env python
__filename = 'gun_support.py'
__fname = 'gun_support'
cStrDividerExcept = '***************************************************************'
cStrDivider = '#================================================================#'
print('', cStrDivider, f'START _ {__filename}', cStrDivider, sep='\n')
print(f'GO {__filename} -> starting IMPORTs and globals decleration')

#------------------------------------------------------------#
#   IMPORTS                                                  #
#------------------------------------------------------------#
import sys, os, re, json, time
import openpyxl, csv
from lxml import html
from datetime import datetime

#------------------------------------------------------------#
#   READ / WRITE SUPPORT                                     #
#------------------------------------------------------------#
def read_js_file(file_path, bPrint=True):
    print('enter _ read_js_file')
    with open(file_path, 'r') as file:
        print('enter with open')
        lines = file.readlines()
        if bPrint: print('lines:\n', *lines, sep='')
        print(f'len(lines): {len(lines)}\n\n')
        print('exit _ read_js_file')
        return lines

def write_lst_dict_to_csv(l : list, file_name : str ='output.csv'):
    print('enter _ write_lst_dict_to_csv')
    #l = [{'name': 'J', 'email': 'j@mail.com'}, {'name': 'K', 'email': 'k@mail.com'}]
    if len(l) == 0: return
    
    field_names = l[0].keys() # column headers
    with open(file_name, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(field_names)
        for d in range(1, len(l)):
            writer.writerow(list(d.values()))
        
def write_lst_to_csv(l : list):
    #l = ['John', 'Doe', 'john.doe@example.com']
    with open('output.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in l:
            writer.writerow(l)
        
def write_dict_to_csv(d : dict):
    #d = {'Name': 'John Doe', 'Email': 'john.doe@example.com'}
    with open('output.csv', 'w', newline='') as csvfile:
        fieldnames = d.keys()
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow(d)

#------------------------------------------------------------#
#   (UN-TESTING) READ / WRITE SUPPORT                        #
#------------------------------------------------------------#
#def write_lst_to_xlsx(l : list):
#    #l = ['John', 'Doe', 'john.doe@example.com']
#    workbook = openpyxl.Workbook()
#    sheet = workbook.active
#    sheet.append(l)
#    workbook.save('output.xlsx')
#
#def write_dict_to_xlsx(d : dict):
#    #d = {'Name': 'John Doe', 'Email': 'john.doe@example.com', 'Age': 30, 'City': 'New York', 'Country': 'USA'}
#    workbook = openpyxl.Workbook()
#    sheet = workbook.active
#
#    # Writing header row
#    header_row = 1
#    for col_num, key in enumerate(d.keys(), start=1):
#        sheet.cell(row=header_row, column=col_num).value = key
#
#    # Writing data rows
#    data_start_row = 2
#    for row_num in range(data_start_row, data_start_row + 5):  # 5 data rows
#        for col_num, value in enumerate(d.values(), start=1):
#            sheet.cell(row=row_num, column=col_num).value = value
#
#    workbook.save('output.xlsx')

#------------------------------------------------------------#
#   DEFAULT SUPPORT                                          #
#------------------------------------------------------------#
def wait_sleep(wait_sec : int, b_print=True): # sleep 'wait_sec'
    print(f'waiting... {wait_sec} sec')
    for s in range(wait_sec, 0, -1):
        if b_print: print('wait ', s, sep='', end='\n')
        time.sleep(1)
    print(f'waited... {wait_sec} sec')
        
def get_time_now():
    return datetime.now().strftime("%H:%M:%S.%f")[0:-4]
    
def read_cli_args():
    print(f'\nread_cli_args...\n # of args: {len(sys.argv)}\n argv lst: {str(sys.argv)}')
    for idx, val in enumerate(sys.argv): print(f' argv[{idx}]: {val}')
    print('read_cli_args _ DONE\n')
    return sys.argv
